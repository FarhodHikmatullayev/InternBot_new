import os
import tempfile
from dataclasses import asdict
from datetime import datetime, timedelta
import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext

from keyboards.default.all_interns import all_interns_in_department_default_keyboard, all_interns_default_keyboard
from keyboards.default.go_to_registration import go_registration_default_keyboard
from keyboards.default.menu_keyboards import back_to_menu, go_back_default_keyboard
from loader import dp, db, bot
from states.download_marks import DownloadMarkState


async def download_marks_of_the_intern_function(profile_id):
    marks = await db.select_marks(intern_id=profile_id)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = 'BAHOLAGAN SHAXS'
    worksheet['C1'] = 'MUOMALASI'
    worksheet['D1'] = 'KIRISHIMLILIGI'
    worksheet['E1'] = 'CHAQQONLIGI VA MALAKASI'
    worksheet['F1'] = "MAS'ULIYATI"
    worksheet['G1'] = "O'ZLASHTIRISH QOBILIYATI"
    worksheet['H1'] = "ICHKI TARTIBGA RIOYASI"
    worksheet['I1'] = "SHAXSIY INTIZOMI"
    worksheet['J1'] = 'BAHOLANGAN VAQT'
    worksheet['K1'] = 'IZOH'

    worksheet.cell(row=1, column=1, value='№')
    worksheet.cell(row=1, column=2, value='BAHOLAGAN SHAXS')
    worksheet.cell(row=1, column=3, value="MUOMALASI")
    worksheet.cell(row=1, column=4, value='KIRISHIMLILIGI')
    worksheet.cell(row=1, column=5, value='CHAQQONLIGI VA MALAKASI')
    worksheet.cell(row=1, column=6, value="MAS'ULIYATI")
    worksheet.cell(row=1, column=7, value="O'ZLASHTIRISH QOBILIYATI")
    worksheet.cell(row=1, column=8, value="ICHKI TARTIBGA RIOYASI")
    worksheet.cell(row=1, column=9, value='SHAXSIY INTIZOMI')
    worksheet.cell(row=1, column=10, value='BAHOLANGAN VAQT')
    worksheet.cell(row=1, column=11, value='IZOH')
    tr = 0
    for row, mark in enumerate(marks, start=2):
        rated_by_id = mark['rated_by_id']
        rated_user = await db.select_user(user_id=rated_by_id)
        rated_user_full_name = rated_user['full_name']

        muomala = mark['muomala']
        kirishimlilik = mark['kirishimlilik']
        chaqqonlik_va_malaka = mark['chaqqonlik_va_malaka']
        masuliyat = mark['masuliyat']
        ozlashtirish_qobiliyati = mark['ozlashtirish_qobiliyati']
        ichki_tartibga_rioyasi = mark['ichki_tartibga_rioyasi']
        shaxsiy_intizomi = mark['shaxsiy_intizomi']
        description = mark['description']
        created_at = mark['created_at']

        tr += 1
        worksheet.cell(row=row, column=1, value=tr)
        worksheet.cell(row=row, column=2, value=rated_user_full_name)
        worksheet.cell(row=row, column=3, value=muomala)
        worksheet.cell(row=row, column=4, value=kirishimlilik)
        worksheet.cell(row=row, column=5, value=chaqqonlik_va_malaka)
        worksheet.cell(row=row, column=6, value=masuliyat)
        worksheet.cell(row=row, column=7, value=ozlashtirish_qobiliyati)
        worksheet.cell(row=row, column=8, value=ichki_tartibga_rioyasi)
        worksheet.cell(row=row, column=9, value=shaxsiy_intizomi)
        worksheet.cell(row=row, column=10,
                       value=(created_at + timedelta(hours=5)).strftime('%d.%m.%Y %H:%M'))
        worksheet.cell(row=row, column=11, value=description)

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'Grades.xlsx')
    workbook.save(file_path)

    return temp_dir


@dp.message_handler(text="📥 Baholarni yuklab olish", state='*')
async def start_download_function(message: types.Message, state: FSMContext):
    try:
        await state.finish()
    except:
        pass
    user_telegram_id = message.from_user.id
    users = await db.select_users(telegram_id=user_telegram_id)
    if not users:
        await message.reply(text="🚫 Sizda botdan foydalanish uchun ruxsat mavjud emas,\n"
                                 "📝 Botdan foydalanish uchun ro'yxatdan o'tishingiz kerak 👇",
                            reply_markup=go_registration_default_keyboard)
    else:
        user = users[0]
        user_role = user['role']
        if user_role not in ['admin', 'hr']:
            await message.reply(text="⚠️ Bu buyruqdan foydalanish uchun sizda ruxsat mavjud emas!",
                                reply_markup=back_to_menu)
        else:
            interns = await db.select_all_intern_profiles()
            if not interns:
                await message.answer(text="⚠️ Hali bazada stajor mavjud emas\n"
                                          "Shuning uchun baholarni yuklab ololmaysiz", reply_markup=back_to_menu)
                return
            markup = await all_interns_default_keyboard()
            await message.answer(text="Stajorlardan birini tanlang 👇", reply_markup=markup)
            await DownloadMarkState.user_id.set()


@dp.message_handler(state=DownloadMarkState.user_id)
async def get_intern_full_name(message: types.Message, state: FSMContext):
    full_name = message.text
    users = await db.select_users(full_name=full_name)
    if not users:
        await message.answer(text="⚠️ Bunday stajor topilmadi", reply_markup=back_to_menu)
        return
    user = users[0]
    user_id = user['id']
    interns = await db.select_intern_profiles(user_id=user_id)
    intern = interns[0]
    await state.update_data(user_id=user_id, profile_id=intern['id'])

    temp_dir = await download_marks_of_the_intern_function(profile_id=intern['id'])

    with open(os.path.join(temp_dir, 'Grades.xlsx'), 'rb') as file:
        await message.answer_document(document=file)

    os.remove(os.path.join(temp_dir, 'Grades.xlsx'))


# download marks for intern
@dp.message_handler(text="📊 Mening baholarim", state="*")
async def download_intern_mark_function(message: types.Message, state: FSMContext):
    try:
        await state.finish()
    except:
        pass
    user_telegram_id = message.from_user.id
    users = await db.select_users(telegram_id=user_telegram_id)
    if not users:
        await message.reply(text="🚫 Sizda botdan foydalanish uchun ruxsat mavjud emas,\n"
                                 "📝 Botdan foydalanish uchun ro'yxatdan o'tishingiz kerak 👇",
                            reply_markup=go_registration_default_keyboard)
    else:
        user = users[0]
        user_role = user['role']
        if user_role != 'intern':
            await message.reply(text="⚠️ Bu buyruqdan foydalanish uchun sizda ruxsat mavjud emas!",
                                reply_markup=back_to_menu)
        else:
            intern_profiles = await db.select_intern_profiles(user_id=user['id'])
            temp_dir = await download_marks_of_the_intern_function(profile_id=intern_profiles[0]['id'])

            with open(os.path.join(temp_dir, 'Grades.xlsx'), 'rb') as file:
                await message.answer_document(document=file)

            os.remove(os.path.join(temp_dir, 'Grades.xlsx'))
